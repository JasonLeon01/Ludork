# -*- encoding: utf-8 -*-

import os
import time
import openpyxl
from typing import Any, Dict, List, Optional, Tuple, cast
from PyQt5 import QtCore, QtGui, QtWidgets
import OfficialLocaleToolsLocaleIO as LocaleIO
from Widgets.SearchLineEdit import AddSearchIcon


class LocaleEditor(QtWidgets.QDialog):

    LOCALE_EXPORTED = QtCore.pyqtSignal()

    _clipboard: Optional[Dict[str, Any]] = None

    def __init__(self, parent, xlsxPath: str):
        super().__init__(parent)
        self._xlsxPath = xlsxPath
        self._modifiedAt: Optional[float] = None
        self._lastSearchQuery = ""
        self._searchMatchIndex = -1
        self._searchMatches: List[Tuple[int, int, int]] = []
        self._wb = openpyxl.load_workbook(xlsxPath)
        self._setupUI()
        self._loadSheets()

    def _setupUI(self) -> None:
        self.setWindowTitle(ELOC("LOCALE_EDITOR_TITLE"))
        self.resize(1000, 640)

        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(8, 8, 8, 8)
        layout.setSpacing(6)

        searchLayout = QtWidgets.QHBoxLayout()
        self._searchEdit = QtWidgets.QLineEdit()
        AddSearchIcon(self._searchEdit)
        self._searchEdit.setPlaceholderText(ELOC("SEARCH"))
        self._searchEdit.setClearButtonEnabled(True)
        self._searchEdit.returnPressed.connect(self._onSearch)
        self._btnSearch = QtWidgets.QPushButton(ELOC("SEARCH"))
        self._btnSearch.clicked.connect(self._onSearch)
        searchLayout.addWidget(self._searchEdit, 1)
        searchLayout.addWidget(self._btnSearch)

        self._tabWidget = QtWidgets.QTabWidget()
        tabBar = cast(QtWidgets.QTabBar, self._tabWidget.tabBar())
        tabBar.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        tabBar.customContextMenuRequested.connect(self._onTabContextMenu)

        btnLayout = QtWidgets.QHBoxLayout()
        self._btnSave = QtWidgets.QPushButton(ELOC("SAVE"))
        self._btnClose = QtWidgets.QPushButton(ELOC("CLOSE"))
        btnLayout.addStretch()
        btnLayout.addWidget(self._btnSave)
        btnLayout.addWidget(self._btnClose)

        self._btnSave.clicked.connect(self._onSave)
        self._btnClose.clicked.connect(self.close)

        layout.addLayout(searchLayout)
        layout.addWidget(self._tabWidget)
        layout.addLayout(btnLayout)

    def _loadSheets(self) -> None:
        self._tabWidget.clear()
        for sheetName in self._wb.sheetnames:
            self._addSheetTab(sheetName)

    def _addSheetTab(self, sheetName: str) -> None:
        ws = self._wb[sheetName]
        rows = ws.max_row or 1
        cols = ws.max_column or 1

        tableRows = max(rows + 5, 10)
        tableCols = max(cols + 2, 5)

        table = QtWidgets.QTableWidget(tableRows, tableCols)
        table.setProperty("sheetName", sheetName)
        table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectItems)
        table.setSelectionMode(QtWidgets.QAbstractItemView.ExtendedSelection)
        cast(QtWidgets.QHeaderView, table.verticalHeader()).setVisible(True)
        table.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        table.customContextMenuRequested.connect(self._onTableContextMenu)

        copyShortcut = QtWidgets.QShortcut(QtGui.QKeySequence.Copy, table, context=QtCore.Qt.WidgetShortcut)
        copyShortcut.activated.connect(lambda t=table: self._onCopy(t))
        pasteShortcut = QtWidgets.QShortcut(QtGui.QKeySequence.Paste, table, context=QtCore.Qt.WidgetShortcut)
        pasteShortcut.activated.connect(lambda t=table: self._onPaste(t))
        deleteShortcut = QtWidgets.QShortcut(QtGui.QKeySequence.Delete, table, context=QtCore.Qt.WidgetShortcut)
        deleteShortcut.activated.connect(lambda t=table: self._onDelete(t))

        table.blockSignals(True)
        for r in range(1, rows + 1):
            for c in range(1, cols + 1):
                cell = ws.cell(row=r, column=c)
                val = cell.value
                item = QtWidgets.QTableWidgetItem(LocaleIO.UnescapeLocaleCellValue(str(val)) if val is not None else "")
                table.setItem(r - 1, c - 1, item)
        table.blockSignals(False)

        table.cellChanged.connect(lambda row, col, t=table, s=sheetName: self._onCellChanged(s, row, col, t))
        self._tabWidget.addTab(table, sheetName)

    def _onTabContextMenu(self, pos: QtCore.QPoint) -> None:
        tabBar = cast(QtWidgets.QTabBar, self._tabWidget.tabBar())
        idx = tabBar.tabAt(pos)
        menu = QtWidgets.QMenu(self)
        actAdd = menu.addAction(ELOC("LOCALE_ADD_SHEET"))
        actRename = None
        actDelete = None
        if idx >= 0:
            actRename = menu.addAction(ELOC("LOCALE_RENAME_SHEET"))
            actDelete = menu.addAction(ELOC("LOCALE_DELETE_SHEET"))
        from Utils import PluginSystem

        PluginSystem.AddRightClickActions(
            menu,
            self,
            "localeSheetTab",
            "hit" if idx >= 0 else "empty",
            self._tabWidget.tabText(idx) if idx >= 0 else None,
        )
        action = menu.exec_(tabBar.mapToGlobal(pos))
        if action == actAdd:
            self._onAddSheet()
        elif actRename is not None and action == actRename:
            self._onRenameSheet(idx)
        elif actDelete is not None and action == actDelete:
            self._onDeleteSheet(idx)

    def _activeTable(self) -> Optional[QtWidgets.QTableWidget]:
        widget = self._tabWidget.currentWidget()
        if isinstance(widget, QtWidgets.QTableWidget):
            return widget
        return None

    def _collectSearchMatches(self, query: str) -> List[Tuple[int, int, int]]:
        normalized = query.casefold()
        matches: List[Tuple[int, int, int]] = []
        for tabIdx in range(self._tabWidget.count()):
            table = self._tabWidget.widget(tabIdx)
            if not isinstance(table, QtWidgets.QTableWidget):
                continue
            for row in range(table.rowCount()):
                for col in range(table.columnCount()):
                    if normalized in self._readCell(table, row, col).casefold():
                        matches.append((tabIdx, row, col))
        return matches

    def _jumpToSearchMatch(self, matchIndex: int) -> None:
        if matchIndex < 0 or matchIndex >= len(self._searchMatches):
            return
        tabIdx, row, col = self._searchMatches[matchIndex]
        self._tabWidget.setCurrentIndex(tabIdx)
        table = self._tabWidget.widget(tabIdx)
        if not isinstance(table, QtWidgets.QTableWidget):
            return
        table.clearSelection()
        table.setCurrentCell(row, col)
        item = table.item(row, col)
        if item is not None:
            table.scrollToItem(item, QtWidgets.QAbstractItemView.PositionAtCenter)
        table.setFocus()

    def _onSearch(self) -> None:
        query = self._searchEdit.text().strip()
        if not query:
            return
        if query != self._lastSearchQuery:
            self._lastSearchQuery = query
            self._searchMatches = self._collectSearchMatches(query)
            self._searchMatchIndex = -1
        if not self._searchMatches:
            return
        self._searchMatchIndex = (self._searchMatchIndex + 1) % len(self._searchMatches)
        self._jumpToSearchMatch(self._searchMatchIndex)

    def _getSelectionRect(self, table: QtWidgets.QTableWidget) -> Optional[Tuple[int, int, int, int]]:
        indexes = table.selectedIndexes()
        if indexes:
            rows = [idx.row() for idx in indexes]
            cols = [idx.column() for idx in indexes]
            return min(rows), min(cols), max(rows), max(cols)
        row = table.currentRow()
        col = table.currentColumn()
        if row >= 0 and col >= 0:
            return row, col, row, col
        return None

    def _readCell(self, table: QtWidgets.QTableWidget, row: int, col: int) -> str:
        item = table.item(row, col)
        return item.text() if item else ""

    def _writeCell(self, sheetName: str, row: int, col: int, val: str) -> None:
        cell = self._wb[sheetName].cell(row=row + 1, column=col + 1)
        cell.value = LocaleIO.EscapeLocaleCellValue(val) if val.strip() else None

    def _syncSheetFromTable(self, table: QtWidgets.QTableWidget) -> None:
        sheetName = table.property("sheetName")
        if not sheetName:
            return
        ws = self._wb[sheetName]
        lastRow = 1
        lastCol = 1
        for row in range(table.rowCount()):
            for col in range(table.columnCount()):
                if self._readCell(table, row, col).strip():
                    lastRow = max(lastRow, row + 1)
                    lastCol = max(lastCol, col + 1)

        maxRow = max(ws.max_row or 1, lastRow)
        maxCol = max(ws.max_column or 1, lastCol)
        for row in range(1, maxRow + 1):
            for col in range(1, maxCol + 1):
                ws.cell(row=row, column=col).value = None
        for row in range(lastRow):
            for col in range(lastCol):
                val = self._readCell(table, row, col)
                ws.cell(row=row + 1, column=col + 1).value = (
                    LocaleIO.EscapeLocaleCellValue(val) if val.strip() else None
                )

        if ws.max_row > lastRow:
            ws.delete_rows(lastRow + 1, ws.max_row - lastRow)
        if ws.max_column > lastCol:
            ws.delete_cols(lastCol + 1, ws.max_column - lastCol)

    def _syncWorkbookFromTables(self) -> None:
        for idx in range(self._tabWidget.count()):
            widget = self._tabWidget.widget(idx)
            if isinstance(widget, QtWidgets.QTableWidget):
                self._syncSheetFromTable(widget)

    def _isFullRowSelection(
        self, table: QtWidgets.QTableWidget, rowStart: int, rowEnd: int, colStart: int, colEnd: int
    ) -> bool:
        if colStart != 0 or colEnd != table.columnCount() - 1:
            return False
        selected = {(idx.row(), idx.column()) for idx in table.selectedIndexes()}
        for row in range(rowStart, rowEnd + 1):
            for col in range(table.columnCount()):
                if (row, col) not in selected:
                    return False
        return True

    def _collectExistingIds(self, table: QtWidgets.QTableWidget) -> set[str]:
        ids: set[str] = set()
        for row in range(1, table.rowCount()):
            item = table.item(row, 0)
            if item:
                text = item.text().strip()
                if text:
                    ids.add(text)
        return ids

    def _makeUniqueId(self, baseId: str, existing: set[str]) -> str:
        baseId = baseId.strip()
        if not baseId:
            return ""
        if baseId not in existing:
            return baseId
        candidate = f"{baseId} (copy)"
        if candidate not in existing:
            return candidate
        index = 2
        while True:
            candidate = f"{baseId} (copy {index})"
            if candidate not in existing:
                return candidate
            index += 1

    def _pasteAnchor(self, table: QtWidgets.QTableWidget) -> Tuple[int, int]:
        rect = self._getSelectionRect(table)
        if rect is not None:
            return rect[0], rect[1]
        return 0, 0

    def _setCellValue(self, table: QtWidgets.QTableWidget, sheetName: str, row: int, col: int, val: str) -> None:
        if row >= table.rowCount():
            table.setRowCount(row + 1)
        if col >= table.columnCount():
            table.setColumnCount(col + 1)
        item = table.item(row, col)
        if item is None:
            item = QtWidgets.QTableWidgetItem(val)
            table.setItem(row, col, item)
        else:
            item.setText(val)
        self._writeCell(sheetName, row, col, val)

    def _onTableContextMenu(self, pos: QtCore.QPoint) -> None:
        table = self.sender()
        if not isinstance(table, QtWidgets.QTableWidget):
            table = self._activeTable()
        if table is None:
            return
        menu = QtWidgets.QMenu(self)
        copyAction = menu.addAction(ELOC("COPY"))
        if copyAction is not None:
            copyAction.setShortcut(QtGui.QKeySequence.Copy)
            copyAction.setEnabled(self._getSelectionRect(table) is not None)
            copyAction.triggered.connect(lambda: self._onCopy(table))
        pasteAction = menu.addAction(ELOC("PASTE"))
        if pasteAction is not None:
            pasteAction.setShortcut(QtGui.QKeySequence.Paste)
            pasteAction.setEnabled(LocaleEditor._clipboard is not None)
            pasteAction.triggered.connect(lambda: self._onPaste(table))
        deleteAction = menu.addAction(ELOC("DELETE"))
        if deleteAction is not None:
            deleteAction.setShortcut(QtGui.QKeySequence.Delete)
            deleteAction.setEnabled(self._getSelectionRect(table) is not None)
            deleteAction.triggered.connect(lambda: self._onDelete(table))
        from Utils import PluginSystem

        PluginSystem.AddRightClickActions(
            menu,
            self,
            "localeTable",
            "always",
            {"sheet": table.property("sheetName"), "selection": self._getSelectionRect(table)},
        )
        menu.exec_(cast(QtWidgets.QWidget, table.viewport()).mapToGlobal(pos))

    def _onCopy(self, table: Optional[QtWidgets.QTableWidget] = None) -> None:
        table = table or self._activeTable()
        if table is None:
            return
        rect = self._getSelectionRect(table)
        if rect is None:
            return
        rowStart, colStart, rowEnd, colEnd = rect
        data: List[List[str]] = []
        for row in range(rowStart, rowEnd + 1):
            data.append([self._readCell(table, row, col) for col in range(colStart, colEnd + 1)])
        LocaleEditor._clipboard = {
            "data": data,
            "fullRows": self._isFullRowSelection(table, rowStart, rowEnd, colStart, colEnd),
        }

    def _onPaste(self, table: Optional[QtWidgets.QTableWidget] = None) -> None:
        clip = LocaleEditor._clipboard
        if not clip:
            return
        table = table or self._activeTable()
        if table is None:
            return
        sheetName = table.property("sheetName")
        if not sheetName:
            return
        data: List[List[str]] = clip.get("data", [])
        if not data:
            return

        anchorRow, anchorCol = self._pasteAnchor(table)
        fullRows = bool(clip.get("fullRows"))
        dedupId = fullRows and anchorCol == 0
        existing: set[str] = set()
        if dedupId:
            existing = self._collectExistingIds(table)
            for row in range(anchorRow, anchorRow + len(data)):
                item = table.item(row, 0)
                if item:
                    text = item.text().strip()
                    if text:
                        existing.discard(text)

        table.blockSignals(True)
        try:
            for rowOffset, rowValues in enumerate(data):
                for colOffset, val in enumerate(rowValues):
                    row = anchorRow + rowOffset
                    col = anchorCol + colOffset
                    if dedupId and col == 0:
                        val = self._makeUniqueId(val, existing)
                        if val:
                            existing.add(val)
                    self._setCellValue(table, sheetName, row, col, val)
        finally:
            table.blockSignals(False)

        pasteRowEnd = anchorRow + len(data) - 1
        pasteColEnd = anchorCol + max(len(row) for row in data) - 1
        table.clearSelection()
        table.setRangeSelected(
            QtWidgets.QTableWidgetSelectionRange(anchorRow, anchorCol, pasteRowEnd, pasteColEnd),
            True,
        )
        table.setCurrentCell(anchorRow, anchorCol)
        self._modifiedAt = time.time()

    def _onDelete(self, table: Optional[QtWidgets.QTableWidget] = None) -> None:
        table = table or self._activeTable()
        if table is None:
            return
        rect = self._getSelectionRect(table)
        if rect is None:
            return
        rowStart, colStart, rowEnd, colEnd = rect
        sheetName = table.property("sheetName")
        if not sheetName:
            return

        table.blockSignals(True)
        try:
            if self._isFullRowSelection(table, rowStart, rowEnd, colStart, colEnd):
                deleteStart = max(rowStart, 1)
                if deleteStart > rowEnd:
                    return
                for row in range(rowEnd, deleteStart - 1, -1):
                    table.removeRow(row)
                if table.rowCount() < 10:
                    table.setRowCount(10)
            else:
                for idx in list(table.selectedIndexes()):
                    self._setCellValue(table, sheetName, idx.row(), idx.column(), "")
        finally:
            table.blockSignals(False)

        self._syncSheetFromTable(table)
        self._modifiedAt = time.time()

    def _collectLanguages(self) -> List[str]:
        langs: List[str] = []
        seen: set[str] = set()
        for sheetName in self._wb.sheetnames:
            ws = self._wb[sheetName]
            headerRow = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not headerRow:
                continue
            headers = ["" if cell is None else str(cell).strip() for cell in headerRow]
            if not headers or headers[0].upper() != "ID":
                continue
            for header in headers[1:]:
                if header and header not in seen:
                    seen.add(header)
                    langs.append(header)
        if langs:
            return langs

        localeDir = os.path.dirname(self._xlsxPath)
        if os.path.isdir(localeDir):
            for entry in sorted(os.listdir(localeDir)):
                path = os.path.join(localeDir, entry)
                if not os.path.isfile(path) or os.path.splitext(entry)[1]:
                    continue
                langs.append(entry)
        if langs:
            return langs

        from Utils import Locale

        editorLangs = Locale.GetLocaleKeys()
        if editorLangs:
            return editorLangs
        return ["en_GB", "zh_CN"]

    def _initSheetHeader(self, ws) -> None:
        headers = ["ID"] + self._collectLanguages()
        for col, value in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=value)

    def _onCellChanged(self, sheetName: str, row: int, col: int, table: QtWidgets.QTableWidget) -> None:
        item = table.item(row, col)
        val = item.text() if item else ""
        self._writeCell(sheetName, row, col, val)
        self._modifiedAt = time.time()

    def _onAddSheet(self) -> None:
        name, ok = QtWidgets.QInputDialog.getText(self, ELOC("LOCALE_ADD_SHEET"), ELOC("LOCALE_SHEET_NAME_PROMPT"))
        if not ok or not name.strip():
            return
        name = name.strip()
        if name in self._wb.sheetnames:
            QtWidgets.QMessageBox.warning(self, "Hint", ELOC("LOCALE_SHEET_EXISTS"))
            return
        ws = self._wb.create_sheet(name)
        self._initSheetHeader(ws)
        self._addSheetTab(name)
        self._modifiedAt = time.time()

    def _onDeleteSheet(self, idx: int) -> None:
        if idx < 0:
            return
        if len(self._wb.sheetnames) <= 1:
            QtWidgets.QMessageBox.warning(self, "Hint", ELOC("LOCALE_CANNOT_DELETE_LAST_SHEET"))
            return
        sheetName = self._tabWidget.tabText(idx)
        res = QtWidgets.QMessageBox.question(
            self,
            "Hint",
            ELOC("LOCALE_CONFIRM_DELETE_SHEET").format(sheetName),
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
            QtWidgets.QMessageBox.No,
        )
        if res != QtWidgets.QMessageBox.Yes:
            return
        del self._wb[sheetName]
        self._tabWidget.removeTab(idx)
        self._modifiedAt = time.time()

    def _onRenameSheet(self, idx: int) -> None:
        if idx < 0:
            return
        oldName = self._tabWidget.tabText(idx)
        newName, ok = QtWidgets.QInputDialog.getText(
            self, ELOC("LOCALE_RENAME_SHEET"), ELOC("LOCALE_SHEET_NAME_PROMPT"), text=oldName
        )
        if not ok or not newName.strip():
            return
        newName = newName.strip()
        if newName == oldName:
            return
        if newName in self._wb.sheetnames:
            QtWidgets.QMessageBox.warning(self, "Hint", ELOC("LOCALE_SHEET_EXISTS"))
            return
        self._wb[oldName].title = newName
        self._tabWidget.setTabText(idx, newName)
        widget = self._tabWidget.widget(idx)
        if isinstance(widget, QtWidgets.QTableWidget):
            widget.setProperty("sheetName", newName)
            widget.cellChanged.disconnect()
            widget.cellChanged.connect(lambda row, col, t=widget, s=newName: self._onCellChanged(s, row, col, t))
        self._modifiedAt = time.time()

    def _onSave(self) -> None:
        localeDir = os.path.dirname(self._xlsxPath)
        try:
            self._syncWorkbookFromTables()
            self._wb.save(self._xlsxPath)
            self._modifiedAt = None
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, "Hint", ELOC("SAVE_FAILED") + "\n" + str(e))
            return
        try:
            if not LocaleIO.ExportLocale(self, self._xlsxPath, localeDir):
                return
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, "Hint", ELOC("EXPORT_LOCALE_FAILED") + "\n" + str(e))
            return
        self.LOCALE_EXPORTED.emit()
