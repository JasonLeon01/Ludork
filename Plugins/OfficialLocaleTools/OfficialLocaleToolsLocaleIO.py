# -*- encoding: utf-8 -*-

from __future__ import annotations

import os
import pickle
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.cell.cell import Cell
from PyQt5 import QtWidgets


_DUPLICATE_ATTRKEY_DISPLAY_LIMIT = 20


def EscapeLocaleCellValue(value: str) -> str:
    if value in ("=", "=="):
        return f"'{value}"
    return value


def UnescapeLocaleCellValue(value: str) -> str:
    if value in ("'=", "'=="):
        return value[1:]
    return value


def _resolveLocaleCellValue(valueCell: Cell, rawCell: Cell) -> Optional[str]:
    val = valueCell.value
    if val is not None:
        return UnescapeLocaleCellValue(str(val))
    if getattr(rawCell, "data_type", None) != "f" or rawCell.value is None:
        return None
    text = str(rawCell.value)
    if text in ("=", "=="):
        return text
    if text.startswith("="):
        text = text[1:]
    text = text.strip()
    return text or None


def _saveData(filePath: str, data: Dict[str, Any]) -> None:
    with open(filePath, "wb") as file:
        pickle.dump(data, file)


def _formatCellLocation(sheetName: str, rowIndex: int) -> str:
    return f"{sheetName}!A{rowIndex}"


def _showDuplicateAttrKeyWarning(parent: Optional[QtWidgets.QWidget], duplicates: List[Tuple[str, str, str]]) -> None:
    lines = [
        f"{key}: {firstLocation}, {duplicateLocation}"
        for key, firstLocation, duplicateLocation in duplicates[:_DUPLICATE_ATTRKEY_DISPLAY_LIMIT]
    ]
    extraCount = len(duplicates) - len(lines)
    if extraCount > 0:
        lines.append(ELOC("LOCALE_DUPLICATE_ATTRKEY_MORE").format(count=extraCount))
    QtWidgets.QMessageBox.warning(
        parent,
        "Hint",
        ELOC("LOCALE_DUPLICATE_ATTRKEY_WARNING").format(duplicates="\n".join(lines)),
    )


def ExportLocale(parent: Optional[QtWidgets.QWidget], xlsxPath: str, localeDir: str) -> bool:
    wbValues = openpyxl.load_workbook(xlsxPath, data_only=True)
    wbCells = openpyxl.load_workbook(xlsxPath, data_only=False)
    langs: List[str] = []
    langMaps: Dict[str, Dict[str, str]] = {}
    keyLocations: Dict[str, str] = {}
    duplicateKeys: List[Tuple[str, str, str]] = []

    for sheetIndex, wsValues in enumerate(wbValues.worksheets):
        wsCells = wbCells.worksheets[sheetIndex]
        headerRow = next(wsValues.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not headerRow:
            continue
        headers = ["" if cell is None else str(cell).strip() for cell in headerRow]
        if not headers or headers[0].upper() != "ID" or len(headers) < 2:
            QtWidgets.QMessageBox.warning(parent, "Hint", ELOC("LOCALE_XLSX_INVALID"))
            return False

        sheetLangs = [h for h in headers[1:] if isinstance(h, str) and h.strip()]
        for lang in sheetLangs:
            if lang not in langMaps:
                langMaps[lang] = {}
                langs.append(lang)

        for rowIndex, rowValues in enumerate(wsValues.iter_rows(min_row=2, values_only=True), start=2):
            if not rowValues:
                continue
            key = rowValues[0]
            if key is None:
                continue
            keyStr = str(key).strip()
            if not keyStr:
                continue
            location = _formatCellLocation(wsValues.title, rowIndex)
            firstLocation = keyLocations.get(keyStr)
            if firstLocation is None:
                keyLocations[keyStr] = location
            else:
                duplicateKeys.append((keyStr, firstLocation, location))
            rawRow = next(wsCells.iter_rows(min_row=rowIndex, max_row=rowIndex), None)
            for i, lang in enumerate(sheetLangs):
                idx = i + 1
                valueCell = wsValues.cell(row=rowIndex, column=idx + 1)
                rawCell = rawRow[idx] if rawRow and idx < len(rawRow) else valueCell
                val = _resolveLocaleCellValue(valueCell, rawCell)
                if val is None:
                    continue
                langMaps[lang][keyStr] = val
    if duplicateKeys:
        _showDuplicateAttrKeyWarning(parent, duplicateKeys)
    for lang, mapping in langMaps.items():
        outPath = os.path.join(localeDir, lang)
        _saveData(outPath, mapping)
    QtWidgets.QMessageBox.information(
        parent,
        "Hint",
        ELOC("EXPORT_LOCALE_SUCCESS").format(langs=", ".join(langs)),
    )
    return True
