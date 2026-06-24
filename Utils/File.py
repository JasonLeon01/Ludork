# -*- encoding: utf-8 -*-

from __future__ import annotations
import os
import sys
import json
import pickle
import shutil
import importlib
import traceback
import openpyxl
from openpyxl.cell.cell import Cell
from pathlib import Path
from typing import Dict, Any, Optional, List, cast
from PyQt5 import QtCore, QtGui, QtWidgets
from EditorGlobal import EditorStatus, MainWindow
from . import System

mainWindow: MainWindow


def GetRootPath() -> str:
    if System.AlreadyPacked():
        return os.path.dirname(sys.executable)
    return os.getcwd()


def GetIniPath() -> str:
    if System.AlreadyPacked() and sys.platform == "darwin":
        path = Path.home() / "Library" / "Application Support" / EditorStatus.APP_NAME
        path.mkdir(parents=True, exist_ok=True)
        return str(path)
    return os.getcwd()


def GetDocPath() -> str:
    from . import Locale

    language = EditorStatus.LANGUAGE if EditorStatus.LANGUAGE in Locale.GetLocaleKeys() else "en_GB"
    return os.path.join(GetRootPath(), "docs", language)


def GetUserPath() -> str:
    if sys.platform == "win32":
        appData = os.getenv("APPDATA")
        if not appData:
            appData = os.path.expanduser(r"~\AppData\Roaming")
        return os.path.join(appData, EditorStatus.APP_NAME)
    elif sys.platform == "darwin":
        path = Path.home() / "Library" / "Application Support" / EditorStatus.APP_NAME
    else:
        print("Unsupported platform")
        sys.exit(1)
    path.mkdir(parents=True, exist_ok=True)
    return str(path)


def GetJSONData(filePath: str) -> Dict[str, Any]:
    with open(filePath, "r", encoding="utf-8") as file:
        jsonData = file.read()
    return json.loads(jsonData)


def SaveJSONData(filePath: str, data: Dict[str, Any]) -> None:
    with open(filePath, "w", encoding="utf-8") as file:
        json.dump(data, file, ensure_ascii=False, indent=4)


def LoadData(filePath: str) -> Any:
    with open(filePath, "rb") as file:
        return pickle.load(file)


def SaveData(filePath, data: Dict[str, Any]) -> None:
    with open(filePath, "wb") as file:
        pickle.dump(data, file)


def _homeDir() -> str:
    try:
        return QtCore.QStandardPaths.writableLocation(QtCore.QStandardPaths.HomeLocation)
    except (RuntimeError, AttributeError):
        return os.path.expanduser("~")


def _getLastPathOrHome() -> str:
    sec = (
        EditorStatus.editorConfig[EditorStatus.APP_NAME]
        if EditorStatus.editorConfig and EditorStatus.APP_NAME in EditorStatus.editorConfig
        else None
    )
    p = sec.get("LastOpenPath") if sec else None
    if isinstance(p, str) and p.strip() and os.path.exists(p):
        return p
    return _homeDir()


def _configPath() -> str:
    return os.path.join(GetIniPath(), f"{EditorStatus.APP_NAME}.ini")


def _setLastOpenPath(path: str) -> None:
    if not EditorStatus.editorConfig:
        return
    if EditorStatus.APP_NAME not in EditorStatus.editorConfig:
        EditorStatus.editorConfig[EditorStatus.APP_NAME] = {}
    EditorStatus.editorConfig[EditorStatus.APP_NAME]["LastOpenPath"] = os.path.abspath(path)
    with open(_configPath(), "w", encoding="utf-8") as f:
        EditorStatus.editorConfig.write(f)


def _openProjectPath(path: str, widget: QtWidgets.QWidget) -> None:
    global mainWindow

    from EditorGlobal import GameData
    from Utils import System

    EditorStatus.PROJ_PATH = os.path.abspath(path)
    if EditorStatus.PROJ_PATH not in sys.path:
        sys.path.insert(0, EditorStatus.PROJ_PATH)
        importlib.invalidate_caches()
        print(f"Add {EditorStatus.PROJ_PATH} to sys.path")
    System.EnsureProjectModulesFresh()
    try:
        GameData.init()
    except Exception as e:
        QtWidgets.QMessageBox.critical(
            None, "Error", ELOC("OPEN_FAILED") + "\n" + str(e) + "\n" + traceback.format_exc()
        )
        sys.exit(1)
    mainWindow = MainWindow(System.GetTitle())
    try:
        cfg_w = int(EditorStatus.editorConfig[EditorStatus.APP_NAME].get("Width", mainWindow.width()))
        cfg_h = int(EditorStatus.editorConfig[EditorStatus.APP_NAME].get("Height", mainWindow.height()))
    except (ValueError, TypeError, KeyError):
        cfg_w, cfg_h = mainWindow.width(), mainWindow.height()
    min_size = mainWindow.minimumSize()
    mainWindow.resize(max(cfg_w, min_size.width()), max(cfg_h, min_size.height()))
    icon_path = os.path.join(GetRootPath(), "Resource", "icon.icns" if sys.platform == "darwin" else "icon.ico")
    if not os.path.exists(icon_path):
        icon_path = os.path.join(GetRootPath(), "Resource", "icon.ico")
    app = cast(Optional[QtWidgets.QApplication], QtWidgets.QApplication.instance())
    if app:
        app.setWindowIcon(QtGui.QIcon(icon_path))
    mainWindow.setWindowIcon(QtGui.QIcon(icon_path))
    screen = app.primaryScreen() if app else None
    fg = mainWindow.frameGeometry()
    cp = screen.availableGeometry().center() if screen else mainWindow.geometry().center()
    fg.moveCenter(cp)
    mainWindow.move(fg.topLeft())
    if app:
        app.aboutToQuit.connect(mainWindow.endGame)
    mainWindow.show()
    widget.close()


def NewProject(parent: QtWidgets.QWidget) -> None:
    root = _getLastPathOrHome()
    dirPath = QtWidgets.QFileDialog.getExistingDirectory(parent, ELOC("SELECT_PROJECT_DIR"), root)
    if not dirPath:
        return
    text, ok = QtWidgets.QInputDialog.getText(
        parent,
        ELOC("ENTER_PROJECT_NAME"),
        ELOC("ENTER_PROJECT_NAME"),
    )
    if not ok:
        return
    name = str(text).strip()
    if not name:
        return
    target = os.path.abspath(os.path.join(dirPath, name))
    if os.path.exists(target):
        QtWidgets.QMessageBox.warning(parent, "Hint", ELOC("PROJECT_EXISTS"))
        return
    try:
        src = os.path.abspath(os.path.join(GetRootPath(), "Sample"))
        if not os.path.isdir(src):
            raise FileNotFoundError(f"Sample directory not found: {src}")
        shutil.copytree(src, target)
        pysfSrc = os.path.abspath(os.path.join(GetRootPath(), "pysf"))
        pysfDst = os.path.join(target, "pysf")
        if os.path.isdir(pysfSrc):
            if os.path.exists(pysfDst) and not os.path.isdir(pysfDst):
                raise FileExistsError(f"Target pysf exists and is not a directory: {pysfDst}")
            if not os.path.exists(pysfDst):
                shutil.copytree(pysfSrc, pysfDst)
        projFile = os.path.join(target, "Main.proj")
        with open(projFile, "w", encoding="utf-8") as f:
            f.write("{}")
    except Exception as e:
        QtWidgets.QMessageBox.critical(None, "Error", ELOC("COPY_FAILED") + "\n" + str(e))
        return
    _setLastOpenPath(target)
    _openProjectPath(target, parent)


def OpenProjectFile(filePath: str, parent: QtWidgets.QWidget) -> bool:
    fp = os.path.abspath(filePath)
    if not fp.lower().endswith(".proj") or not os.path.isfile(fp):
        QtWidgets.QMessageBox.warning(parent, "Hint", ELOC("INVALID_PROJ_FILE"))
        return False
    proj_dir = os.path.dirname(fp)
    _setLastOpenPath(proj_dir)
    _openProjectPath(proj_dir, parent)
    return True


def OpenProject(parent: QtWidgets.QWidget) -> None:
    root = _getLastPathOrHome()
    fp, _ = QtWidgets.QFileDialog.getOpenFileName(
        parent,
        ELOC("SELECT_PROJ_FILE"),
        root,
        "Project Files (*.proj)",
    )
    if not fp:
        return
    OpenProjectFile(fp, parent)


def EscapeLocaleCellValue(value: str) -> str:
    if value in ("=", "=="):
        return f"'{value}"
    return value


def UnescapeLocaleCellValue(value: str) -> str:
    if value in ("'=", "'=="):
        return value[1:]
    return value


def _resolveLocaleCellValue(valueCell: Cell, rawCell: Cell) -> Optional[str]:
    val = valueCell.value
    if val is not None:
        return UnescapeLocaleCellValue(str(val))
    if getattr(rawCell, "data_type", None) != "f" or rawCell.value is None:
        return None
    text = str(rawCell.value)
    if text in ("=", "=="):
        return text
    if text.startswith("="):
        text = text[1:]
    text = text.strip()
    return text or None


def ExportLocale(parent: Optional[QtWidgets.QWidget], xlsxPath: str, localeDir: str) -> None:
    wbValues = openpyxl.load_workbook(xlsxPath, data_only=True)
    wbCells = openpyxl.load_workbook(xlsxPath, data_only=False)
    langs: List[str] = []
    langMaps: Dict[str, Dict[str, str]] = {}

    for sheetIndex, wsValues in enumerate(wbValues.worksheets):
        wsCells = wbCells.worksheets[sheetIndex]
        headerRow = next(wsValues.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not headerRow:
            continue
        headers = ["" if cell is None else str(cell).strip() for cell in headerRow]
        if not headers or headers[0].upper() != "ID" or len(headers) < 2:
            QtWidgets.QMessageBox.warning(parent, "Hint", ELOC("LOCALE_XLSX_INVALID"))
            return

        sheetLangs = [h for h in headers[1:] if isinstance(h, str) and h.strip()]
        for lang in sheetLangs:
            if lang not in langMaps:
                langMaps[lang] = {}
                langs.append(lang)

        for rowIndex, rowValues in enumerate(wsValues.iter_rows(min_row=2, values_only=True), start=2):
            if not rowValues:
                continue
            key = rowValues[0]
            if key is None:
                continue
            keyStr = str(key).strip()
            if not keyStr:
                continue
            rawRow = next(wsCells.iter_rows(min_row=rowIndex, max_row=rowIndex), None)
            for i, lang in enumerate(sheetLangs):
                idx = i + 1
                valueCell = wsValues.cell(row=rowIndex, column=idx + 1)
                rawCell = rawRow[idx] if rawRow and idx < len(rawRow) else valueCell
                val = _resolveLocaleCellValue(valueCell, rawCell)
                if val is None:
                    continue
                langMaps[lang][keyStr] = val
    for lang, mapping in langMaps.items():
        outPath = os.path.join(localeDir, lang)
        SaveData(outPath, mapping)
    QtWidgets.QMessageBox.information(
        parent,
        "Hint",
        ELOC("EXPORT_LOCALE_SUCCESS").format(langs=", ".join(langs)),
    )
